from counter_party.helpers import counter_party_balance_income
from counter_party.models import CounterParty
from .models import ProjectSetting, PaymentLog
from order.models import RetailOrder, Order, ReturnItem
from income.models import Income
from datetime import datetime
from django.db.models import Sum, F, Subquery, OuterRef
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.writer.excel import save_virtual_workbook
import openpyxl, os
from django.http import HttpResponse


def payment_income_action(outcat, model_id, outlay_id, payment_method, amount, user, **kwargs):
    PaymentLog.objects.create(
        user=user,
        outcat=outcat,
        model_id=model_id,
        outlay=outlay_id,
        payment_type='income',
        payment_method=payment_method,
        amount=amount,
        status='accepted',
    )
    return 1


def payment_outcome_action(outcat, model_id, outlay_id, payment_method, amount, user, **kwargs):
    PaymentLog.objects.create(
        user=user,
        outcat=outcat,
        model_id=model_id,
        outlay=outlay_id,
        payment_type='outcome',
        payment_method=payment_method,
        amount=amount,
        status='accepted',
    )
    return 1


def payment_accept(request, kwargs):
    payment = PaymentLog.objects.get(id=request.POST.get('payment_id'))
    payment.status = 'accepted'
    counter_party_balance_income(payment.model_id, payment.amount)
    payment.save()
    return 1


def payment_reject(request, kwargs):
    payment = PaymentLog.objects.get(id=request.POST.get('payment_id'))
    payment.status = 'rejected'
    counter_party_balance_income(payment.model_id, payment.amount)
    payment.save()
    return 1


def payment_theory_excel(request):
    retail_orders = RetailOrder.objects.filter(status='completed')

    orders = Order.objects.filter(status__in=['delivered', 'completed'])

    incomes = Income.objects.filter(status='completed')

    return_items = ReturnItem.objects.filter(status='write-off')

    start_date = request.GET.get('start_date') if request.GET.get('start_date') != '' else None
    end_date = request.GET.get('end_date') if request.GET.get('start_date') != '' else None

    if start_date and end_date:
        start_date += ' 00:00:00'
        end_date += ' 23:59:59'
        retail_orders = retail_orders.filter(
            created__range=[
                start_date,
                end_date,
            ]
        )
        orders = orders.filter(
            created__range=[
                start_date,
                end_date,
            ]
        )
        incomes = incomes.filter(
            created__range=[
                start_date,
                end_date,
            ]
        )
        return_items = return_items.filter(
            created__range=[
                start_date,
                end_date,
            ]
        )
    else:
        today_start = datetime.today().date().strftime('%Y-%m-%d') + ' 00:00:00'
        today_end = datetime.today().date().strftime('%Y-%m-%d') + ' 23:59:59'
        retail_orders = retail_orders.filter(
            created__range=[
                today_start,
                today_end,
            ]
        )
        orders = orders.filter(
            created__range=[
                today_start,
                today_end,
            ]
        )
        incomes = incomes.filter(
            created__range=[
                today_start,
                today_end,
            ]
        )
        return_items = return_items.filter(
            created__range=[
                today_start,
                today_end,
            ]
        )

    orders = orders.annotate(
        deliver_percent=F('invoice__total') * (F('invoice__deliver__service_percent') / 100),
        agent_percent=F('invoice__total') * (F('agent__service_percent') / 100),
        self_price=Sum(F('order_items__warehouse_product__self_price') * F('order_items__count')),
        profit_amount=F('invoice__total') - F('self_price') - F('deliver_percent') - F('agent_percent'),
    )

    book = openpyxl.load_workbook('payment_report.xlsx')
    booksheet = book.worksheets[0]

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    bold_font_bold = Font(bold=True)
    cell_color = PatternFill(fill_type='solid', start_color="b3b6ba", end_color="b3b6ba")

    booksheet['c2'] = "Отчет по прибыли (Теоретическая)"
    booksheet['c2'].font = bold_font_bold

    booksheet['b4'] = "Доход(От розничной торговли) на Сегодня"
    booksheet['b4'].font = bold_font_bold

    bold_font = Font(bold=True, size=8)

    booksheet['a6'] = "ЗАКАЗ №"
    booksheet['a6'].font = bold_font
    booksheet['a6'].border = border
    booksheet['b6'] = "ДАТА СОЗДАНИЕ"
    booksheet['b6'].font = bold_font
    booksheet['b6'].border = border
    booksheet['c6'] = "КЛИЕНТ"
    booksheet['c6'].font = bold_font
    booksheet['c6'].border = border
    booksheet['d6'] = "ПРИНЯЛ"
    booksheet['d6'].font = bold_font
    booksheet['d6'].border = border
    booksheet['e6'] = "СТАТУС"
    booksheet['e6'].font = bold_font
    booksheet['e6'].border = border
    booksheet['f6'] = "СУММА"
    booksheet['f6'].font = bold_font
    booksheet['f6'].border = border

    counter = 0
    for retail_order in retail_orders:
        counter += 1
        booksheet[f'A{6 + counter}'] = f"{retail_order.id}"
        booksheet[f'A{6 + counter}'].border = border
        booksheet[f'B{6 + counter}'] = f"{retail_order.created.strftime('%d.%m.%Y %H:%M')}"
        booksheet[f'B{6 + counter}'].border = border
        booksheet[f'C{6 + counter}'] = f"{retail_order.client}"
        booksheet[f'C{6 + counter}'].border = border
        if retail_order.user.fullname:
            booksheet[f'D{6 + counter}'] = f"{retail_order.user.fullname}"
        else:
            booksheet[f'D{6 + counter}'] = f"{retail_order.user.username}"
        booksheet[f'D{6 + counter}'].border = border
        booksheet[f'E{6 + counter}'] = f"{retail_order.get_status_display()}"
        booksheet[f'E{6 + counter}'].border = border
        booksheet[f'F{6 + counter}'] = f"{retail_order.total}"
        booksheet[f'F{6 + counter}'].border = border
    if not retail_orders:
        booksheet[f'A7'].border = border
        booksheet[f'B7'].border = border
        booksheet[f'C7'] = 'Пусто'
        booksheet[f'C7'].border = border
        booksheet[f'D7'].border = border
        booksheet[f'E7'].border = border
        booksheet[f'F7'].border = border

    booksheet[f'b{8 + counter}'] = "Доход(От торговли) на Сегодня"
    booksheet[f'b{8 + counter}'].font = bold_font_bold

    booksheet[f'a{10 + counter}'] = "#"
    booksheet[f'a{10 + counter}'].font = bold_font
    booksheet[f'a{10 + counter}'].border = border
    booksheet[f'b{10 + counter}'] = "ДАТА"
    booksheet[f'b{10 + counter}'].font = bold_font
    booksheet[f'b{10 + counter}'].border = border
    booksheet[f'c{10 + counter}'] = "ЗАКАЗ"
    booksheet[f'c{10 + counter}'].font = bold_font
    booksheet[f'c{10 + counter}'].border = border
    booksheet[f'd{10 + counter}'] = "СТАТУС"
    booksheet[f'd{10 + counter}'].font = bold_font
    booksheet[f'd{10 + counter}'].border = border
    booksheet[f'e{10 + counter}'] = "СУММА"
    booksheet[f'e{10 + counter}'].font = bold_font
    booksheet[f'e{10 + counter}'].border = border
    booksheet[f'f{10 + counter}'] = "СЕБЕС."
    booksheet[f'f{10 + counter}'].font = bold_font
    booksheet[f'f{10 + counter}'].border = border
    booksheet[f'g{10 + counter}'] = "ПРОЦ. ДОСТАВ."
    booksheet[f'g{10 + counter}'].font = bold_font
    booksheet[f'g{10 + counter}'].border = border
    booksheet[f'h{10 + counter}'] = "ПРОЦ. АГЕНТА."
    booksheet[f'h{10 + counter}'].font = bold_font
    booksheet[f'h{10 + counter}'].border = border
    booksheet[f'i{10 + counter}'] = "ПРИБЫЛЬ"
    booksheet[f'i{10 + counter}'].font = bold_font
    booksheet[f'i{10 + counter}'].border = border

    counter_1 = 0
    for order in orders:
        counter_1 += 1
        booksheet[f'A{10 + counter + counter_1}'] = f"{order.id}"
        booksheet[f'A{10 + counter + counter_1}'].border = border
        booksheet[f'B{10 + counter + counter_1}'] = f"{order.created.strftime('%d.%m.%Y %H:%M')}"
        booksheet[f'B{10 + counter + counter_1}'].border = border
        booksheet[f'C{10 + counter + counter_1}'] = f"{order.counter_party}"
        booksheet[f'C{10 + counter + counter_1}'].border = border
        booksheet[f'D{10 + counter + counter_1}'] = f"{order.get_status_display()}"
        booksheet[f'D{10 + counter + counter_1}'].border = border
        for invoice in order.invoice_set.all():
            if invoice.total:
                booksheet[f'E{10 + counter + counter_1}'] = f"{invoice.total} сум"
            else:
                booksheet[f'E{10 + counter + counter_1}'] = "0 сум"
        booksheet[f'E{10 + counter + counter_1}'].border = border
        if order.self_price:
            booksheet[f'F{10 + counter + counter_1}'] = f"{order.self_price} сум"
        else:
            booksheet[f'F{10 + counter + counter_1}'] = f"0 сум"
        booksheet[f'F{10 + counter + counter_1}'].border = border
        if order.deliver_percent:
            booksheet[f'G{10 + counter + counter_1}'] = f"{order.deliver_percent} сум"
        else:
            booksheet[f'G{10 + counter + counter_1}'] = f"0 сум"
        booksheet[f'G{10 + counter + counter_1}'].border = border
        if order.agent_percent:
            booksheet[f'H{10 + counter + counter_1}'] = f"{order.agent_percent} сум"
        else:
            booksheet[f'H{10 + counter + counter_1}'] = f"0 сум"
        booksheet[f'H{10 + counter + counter_1}'].border = border
        if order.profit_amount:
            booksheet[f'I{10 + counter + counter_1}'] = f"{order.profit_amount} сум"
        else:
            booksheet[f'I{10 + counter + counter_1}'] = f"0 сум"
        booksheet[f'I{10 + counter + counter_1}'].border = border
    if not orders:
        booksheet[f'A{11 + counter + counter_1}'].border = border
        booksheet[f'B{11 + counter + counter_1}'].border = border
        booksheet[f'C{11 + counter + counter_1}'] = 'Пусто'
        booksheet[f'C{11 + counter + counter_1}'].border = border
        booksheet[f'D{11 + counter + counter_1}'].border = border
        booksheet[f'E{11 + counter + counter_1}'].border = border
        booksheet[f'F{11 + counter + counter_1}'].border = border
        booksheet[f'G{11 + counter + counter_1}'].border = border
        booksheet[f'H{11 + counter + counter_1}'].border = border
        booksheet[f'I{11 + counter + counter_1}'].border = border

    booksheet[f'b{13 + counter + counter_1}'] = "Расход(На покупку товаров) на Сегодня"
    booksheet[f'b{13 + counter + counter_1}'].font = bold_font_bold

    booksheet[f'a{15 + counter + counter_1}'] = "ПРИХОД №"
    booksheet[f'a{15 + counter + counter_1}'].font = bold_font
    booksheet[f'a{15 + counter + counter_1}'].border = border
    booksheet[f'b{15 + counter + counter_1}'] = "ДАТА СОЗДАНИЕ"
    booksheet[f'b{15 + counter + counter_1}'].font = bold_font
    booksheet[f'b{15 + counter + counter_1}'].border = border
    booksheet[f'c{15 + counter + counter_1}'] = "ПРИНЯЛ"
    booksheet[f'c{15 + counter + counter_1}'].font = bold_font
    booksheet[f'c{15 + counter + counter_1}'].border = border
    booksheet[f'd{15 + counter + counter_1}'] = "ПОСТАВЩИК"
    booksheet[f'd{15 + counter + counter_1}'].font = bold_font
    booksheet[f'd{15 + counter + counter_1}'].border = border
    booksheet[f'e{15 + counter + counter_1}'] = "СТАТУС"
    booksheet[f'e{15 + counter + counter_1}'].font = bold_font
    booksheet[f'e{15 + counter + counter_1}'].border = border
    booksheet[f'f{15 + counter + counter_1}'] = "СУММА."
    booksheet[f'f{15 + counter + counter_1}'].font = bold_font
    booksheet[f'f{15 + counter + counter_1}'].border = border

    counter_2 = 0
    for income in incomes:
        counter_2 += 1
        booksheet[f'A{15 + counter + counter_1 + counter_2}'] = f"{income.id}"
        booksheet[f'A{15 + counter + counter_1 + counter_2}'].border = border
        booksheet[f'B{15 + counter + counter_1 + counter_2}'] = f"{income.created.strftime('%d.%m.%Y %H:%M')}"
        booksheet[f'B{15 + counter + counter_1 + counter_2}'].border = border
        booksheet[
            f'C{15 + counter + counter_1 + counter_2}'] = f"{income.user.username} | {income.user.get_user_type_display()}"
        booksheet[f'C{15 + counter + counter_1 + counter_2}'].border = border
        booksheet[f'D{15 + counter + counter_1 + counter_2}'] = f"{income.provider}"
        booksheet[f'D{15 + counter + counter_1 + counter_2}'].border = border
        booksheet[f'E{15 + counter + counter_1 + counter_2}'] = f"{income.get_status_display()} сум"
        booksheet[f'E{15 + counter + counter_1 + counter_2}'].border = border
        booksheet[f'F{15 + counter + counter_1 + counter_2}'] = f"{income.total} сум"
        booksheet[f'F{15 + counter + counter_1 + counter_2}'].border = border
    if not incomes:
        booksheet[f'A{16 + counter + counter_1 + counter_2}'].border = border
        booksheet[f'B{16 + counter + counter_1 + counter_2}'].border = border
        booksheet[f'C{16 + counter + counter_1 + counter_2}'] = 'Пусто'
        booksheet[f'C{16 + counter + counter_1 + counter_2}'].border = border
        booksheet[f'D{16 + counter + counter_1 + counter_2}'].border = border
        booksheet[f'E{16 + counter + counter_1 + counter_2}'].border = border
        booksheet[f'F{16 + counter + counter_1 + counter_2}'].border = border

    booksheet[f'b{18 + counter + counter_1 + counter_2}'] = "Расход(На возврат товаров) на Сегодня"
    booksheet[f'b{18 + counter + counter_1 + counter_2}'].font = bold_font_bold

    booksheet[f'a{20 + counter + counter_1 + counter_2}'] = "ПРИХОД №"
    booksheet[f'a{20 + counter + counter_1 + counter_2}'].font = bold_font
    booksheet[f'a{20 + counter + counter_1 + counter_2}'].border = border
    booksheet[f'b{20 + counter + counter_1 + counter_2}'] = "ДАТА СОЗДАНИЕ"
    booksheet[f'b{20 + counter + counter_1 + counter_2}'].font = bold_font
    booksheet[f'b{20 + counter + counter_1 + counter_2}'].border = border
    booksheet[f'c{20 + counter + counter_1 + counter_2}'] = "ПРИНЯЛ"
    booksheet[f'c{20 + counter + counter_1 + counter_2}'].font = bold_font
    booksheet[f'c{20 + counter + counter_1 + counter_2}'].border = border
    booksheet[f'd{20 + counter + counter_1 + counter_2}'] = "КЛИЕНТ"
    booksheet[f'd{20 + counter + counter_1 + counter_2}'].font = bold_font
    booksheet[f'd{20 + counter + counter_1 + counter_2}'].border = border
    booksheet[f'e{20 + counter + counter_1 + counter_2}'] = "СТАТУС"
    booksheet[f'e{20 + counter + counter_1 + counter_2}'].font = bold_font
    booksheet[f'e{20 + counter + counter_1 + counter_2}'].border = border
    booksheet[f'f{20 + counter + counter_1 + counter_2}'] = "СУММА."
    booksheet[f'f{20 + counter + counter_1 + counter_2}'].font = bold_font
    booksheet[f'f{20 + counter + counter_1 + counter_2}'].border = border

    counter_3 = 0
    for return_item in return_items:
        counter_3 += 1
        booksheet[f'A{20 + counter + counter_1 + counter_2 + counter_3}'] = f"{return_item.id}"
        booksheet[f'A{20 + counter + counter_1 + counter_2 + counter_3}'].border = border
        booksheet[
            f'B{20 + counter + counter_1 + counter_2 + counter_3}'] = f"{return_item.created.strftime('%d.%m.%Y %H:%M')}"
        booksheet[f'B{20 + counter + counter_1 + counter_2 + counter_3}'].border = border
        booksheet[
            f'C{20 + counter + counter_1 + counter_2 + counter_3}'] = f"{return_item.deliver.user.username} | {return_item.deliver.user.get_user_type_display()}"
        booksheet[f'C{20 + counter + counter_1 + counter_2 + counter_3}'].border = border
        booksheet[f'D{20 + counter + counter_1 + counter_2 + counter_3}'] = f"{return_item.counter_party.full_name}"
        booksheet[f'D{20 + counter + counter_1 + counter_2 + counter_3}'].border = border
        booksheet[f'E{20 + counter + counter_1 + counter_2 + counter_3}'] = f"{return_item.get_status_display()} сум"
        booksheet[f'E{20 + counter + counter_1 + counter_2 + counter_3}'].border = border
        booksheet[f'F{20 + counter + counter_1 + counter_2 + counter_3}'] = f"{return_item.total} сум"
        booksheet[f'F{20 + counter + counter_1 + counter_2 + counter_3}'].border = border
    if not return_items:
        booksheet[f'A{21 + counter + counter_1 + counter_2 + counter_3}'].border = border
        booksheet[f'B{21 + counter + counter_1 + counter_2 + counter_3}'].border = border
        booksheet[f'C{21 + counter + counter_1 + counter_2 + counter_3}'] = 'Пусто'
        booksheet[f'C{21 + counter + counter_1 + counter_2 + counter_3}'].border = border
        booksheet[f'D{21 + counter + counter_1 + counter_2 + counter_3}'].border = border
        booksheet[f'E{21 + counter + counter_1 + counter_2 + counter_3}'].border = border
        booksheet[f'F{21 + counter + counter_1 + counter_2 + counter_3}'].border = border

    retail_orders = retail_orders.aggregate(total=Sum('total'))['total']
    orders = orders.aggregate(total=Sum('profit_amount'))['total']
    incomes = incomes.aggregate(total=Sum('total'))['total']
    return_items = return_items.aggregate(total=Sum('total'))['total']

    if retail_orders is None:
        retail_orders = 0
    if orders is None:
        orders = 0
    if incomes is None:
        incomes = 0
    if return_items is None:
        return_items = 0

    total_income = retail_orders + orders
    total_outcome = incomes + return_items
    real_income = (orders + retail_orders) - incomes - return_items

    if total_income == 0:
        booksheet[f'b{24 + counter + counter_1 + counter_2 + counter_3}'] = f"Общий Доход: 0 сум"
    else:
        booksheet[f'b{24 + counter + counter_1 + counter_2 + counter_3}'] = f"Общий Доход: {total_income} сум"
    booksheet[f'b{24 + counter + counter_1 + counter_2 + counter_3}'].font = bold_font_bold
    if total_outcome == 0:
        booksheet[f'c{24 + counter + counter_1 + counter_2 + counter_3}'] = f"Общий Расход: 0 сум"
    else:
        booksheet[f'c{24 + counter + counter_1 + counter_2 + counter_3}'] = f"Общий Расход: {total_outcome} сум"
    booksheet[f'c{24 + counter + counter_1 + counter_2 + counter_3}'].font = bold_font_bold
    if real_income == 0:
        booksheet[f'd{24 + counter + counter_1 + counter_2 + counter_3}'] = f"Общий Прибыл: 0 сум"
    else:
        booksheet[f'd{24 + counter + counter_1 + counter_2 + counter_3}'] = f"Общий Прибыл: {real_income} сум"
    booksheet[f'd{24 + counter + counter_1 + counter_2 + counter_3}'].font = bold_font_bold

    response = HttpResponse(content=save_virtual_workbook(book),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="payment_theory.xlsx"'

    return response


def payment_real_excel(request):
    payments_income = PaymentLog.objects.filter(status='accepted', payment_type='income')
    payments_outcome = PaymentLog.objects.filter(status='accepted', payment_type='outcome')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    payments_income = payments_income.annotate(
        counter_party_name=Subquery(CounterParty.objects
                                    .filter(id=OuterRef('model_id'))
                                    .values_list('full_name'))
    )
    payments_outcome = payments_outcome.annotate(
        counter_party_name=Subquery(CounterParty.objects
                                    .filter(id=OuterRef('model_id'))
                                    .values_list('full_name'))
    )

    if start_date and end_date:
        start_date += ' 00:00:00'
        end_date += ' 23:59:59'
        payments_income = payments_income.filter(
            created__range=[
                start_date,
                end_date,
            ]
        )
        payments_outcome = payments_outcome.filter(
            created__range=[
                start_date,
                end_date,
            ]
        )
    else:
        today_start = datetime.today().date().strftime('%Y-%m-%d') + ' 00:00:00'
        today_end = datetime.today().date().strftime('%Y-%m-%d') + ' 23:59:59'
        payments_income = payments_income.filter(
            created__range=[
                today_start,
                today_end,
            ]
        )
        payments_outcome = payments_outcome.filter(
            created__range=[
                today_start,
                today_end,
            ]
        )
    payments_outcome = payments_outcome
    payments_outcome_total = payments_outcome.aggregate(total=Sum('amount'))['total']
    payments_income = payments_income
    payments_income_total = payments_income.aggregate(total=Sum('amount'))['total']

    book = openpyxl.load_workbook('payment_report.xlsx')
    booksheet = book.worksheets[0]

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    bold_font_bold = Font(bold=True)
    cell_color = PatternFill(fill_type='solid', start_color="b3b6ba", end_color="b3b6ba")

    booksheet['c2'] = "Отчет по прибыли (Реальное)"
    booksheet['c2'].font = bold_font_bold

    booksheet['b4'] = "Доход на Сегодня"
    booksheet['b4'].font = bold_font_bold

    bold_font = Font(bold=True, size=8)

    booksheet['a6'] = "#"
    booksheet['a6'].font = bold_font
    booksheet['a6'].border = border
    booksheet['b6'] = "ДАТА СОЗДАНИЕ"
    booksheet['b6'].font = bold_font
    booksheet['b6'].border = border
    booksheet['c6'] = "ПРИНЯЛ"
    booksheet['c6'].font = bold_font
    booksheet['c6'].border = border
    booksheet['d6'] = "КОНТРАГЕНТ"
    booksheet['d6'].font = bold_font
    booksheet['d6'].border = border
    booksheet['e6'] = "СУММА"
    booksheet['e6'].font = bold_font
    booksheet['e6'].border = border
    booksheet['f6'] = "СТАТУС"
    booksheet['f6'].font = bold_font
    booksheet['f6'].border = border

    counter = 0
    for payment in payments_income:
        counter += 1
        booksheet[f'A{6 + counter}'] = f"{payment.id}"
        booksheet[f'A{6 + counter}'].border = border
        booksheet[f'B{6 + counter}'] = f"{payment.created.strftime('%d.%m.%Y %H:%M')}"
        booksheet[f'B{6 + counter}'].border = border
        booksheet[f'C{6 + counter}'] = f"{payment.user.username} | {payment.user.get_user_type_display()}"
        booksheet[f'C{6 + counter}'].border = border
        if payment.counter_party_name:
            booksheet[f'D{6 + counter}'] = f"{payment.counter_party_name}"
        else:
            booksheet[f'D{6 + counter}'] = "-"
        booksheet[f'D{6 + counter}'].border = border
        booksheet[f'E{6 + counter}'] = f"{payment.amount}"
        booksheet[f'E{6 + counter}'].border = border
        booksheet[f'F{6 + counter}'] = f"{payment.get_status_display()}"
        booksheet[f'F{6 + counter}'].border = border
    if not payments_income:
        booksheet[f'A7'].border = border
        booksheet[f'B7'].border = border
        booksheet[f'C7'] = 'Пусто'
        booksheet[f'C7'].border = border
        booksheet[f'D7'].border = border
        booksheet[f'E7'].border = border
        booksheet[f'F7'].border = border

    booksheet[f'b{8 + counter}'] = "Расход на Сегодня"
    booksheet[f'b{8 + counter}'].font = bold_font_bold

    booksheet[f'a{10 + counter}'] = "#"
    booksheet[f'a{10 + counter}'].font = bold_font
    booksheet[f'a{10 + counter}'].border = border
    booksheet[f'b{10 + counter}'] = "ДАТА СОЗДАНИЕ"
    booksheet[f'b{10 + counter}'].font = bold_font
    booksheet[f'b{10 + counter}'].border = border
    booksheet[f'c{10 + counter}'] = "ПРИНЯЛ"
    booksheet[f'c{10 + counter}'].font = bold_font
    booksheet[f'c{10 + counter}'].border = border
    booksheet[f'd{10 + counter}'] = "КОНТРАГЕНТ"
    booksheet[f'd{10 + counter}'].font = bold_font
    booksheet[f'd{10 + counter}'].border = border
    booksheet[f'e{10 + counter}'] = "СУММА"
    booksheet[f'e{10 + counter}'].font = bold_font
    booksheet[f'e{10 + counter}'].border = border
    booksheet[f'f{10 + counter}'] = "СТАТУС"
    booksheet[f'f{10 + counter}'].font = bold_font
    booksheet[f'f{10 + counter}'].border = border

    counter_1 = 0
    for payment in payments_outcome:
        counter_1 += 1
        booksheet[f'A{10 + counter + counter_1}'] = f"{payment.id}"
        booksheet[f'A{10 + counter + counter_1}'].border = border
        booksheet[f'B{10 + counter + counter_1}'] = f"{payment.created.strftime('%d.%m.%Y %H:%M')}"
        booksheet[f'B{10 + counter + counter_1}'].border = border
        booksheet[f'C{10 + counter + counter_1}'] = f"{payment.user.username} | {payment.user.get_user_type_display()}"
        booksheet[f'C{10 + counter + counter_1}'].border = border
        if payment.counter_party_name:
            booksheet[f'D{10 + counter + counter_1}'] = f"{payment.counter_party_name}"
        else:
            booksheet[f'D{10 + counter + counter_1}'] = "-"
        booksheet[f'D{10 + counter + counter_1}'].border = border
        booksheet[f'E{10 + counter + counter_1}'] = f"{payment.amount} сум"
        booksheet[f'E{10 + counter + counter_1}'].border = border
        booksheet[f'F{10 + counter + counter_1}'] = f"{payment.get_status_display()}"
        booksheet[f'F{10 + counter + counter_1}'].border = border
    if not payments_outcome:
        booksheet[f'A{11 + counter + counter_1}'].border = border
        booksheet[f'B{11 + counter + counter_1}'].border = border
        booksheet[f'C{11 + counter + counter_1}'] = 'Пусто'
        booksheet[f'C{11 + counter + counter_1}'].border = border
        booksheet[f'D{11 + counter + counter_1}'].border = border
        booksheet[f'E{11 + counter + counter_1}'].border = border
        booksheet[f'F{11 + counter + counter_1}'].border = border

    if payments_outcome_total is None:
        payments_outcome_total = 0
    if payments_income_total is None:
        payments_income_total = 0

    total_income = payments_income_total
    total_outcome = payments_outcome_total
    real_income = payments_income_total - payments_outcome_total

    if total_income == 0:
        booksheet[f'b{13 + counter + counter_1}'] = f"Общий Доход: 0 сум"
    else:
        booksheet[f'b{13 + counter + counter_1}'] = f"Общий Доход: {total_income} сум"
    booksheet[f'b{13 + counter + counter_1}'].font = bold_font_bold
    if total_outcome == 0:
        booksheet[f'c{13 + counter + counter_1}'] = f"Общий Расход: 0 сум"
    else:
        booksheet[f'c{13 + counter + counter_1}'] = f"Общий Расход: {total_outcome} сум"
    booksheet[f'c{13 + counter + counter_1}'].font = bold_font_bold
    if real_income == 0:
        booksheet[f'd{13 + counter + counter_1}'] = f"Общий Прибыл: 0 сум"
    else:
        booksheet[f'd{13 + counter + counter_1}'] = f"Общий Прибыл: {real_income} сум"
    booksheet[f'd{13 + counter + counter_1}'].font = bold_font_bold

    response = HttpResponse(content=save_virtual_workbook(book),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="payment_theory.xlsx"'

    return response

import locale

from counter_party.helpers import counter_party_balance_outcome
from payment.helpers import payment_income_action, payment_outcome_action
from payment.models import Outlay
from .models import Order, OrderItem, Invoice, InvoiceItem, InvoiceProvider, ReturnItem, RetailOrderItem, RetailOrder
from django.db.models import Sum
from django.http import HttpResponse
from django.conf import settings
from warehouse.models import WarehouseProduct
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.writer.excel import save_virtual_workbook
import openpyxl, os

locale.setlocale(locale.LC_NUMERIC, 'ru_RU.utf8')


def add_order_item(request, kwargs):
    order_id = kwargs['pk']
    warehouse_product_id = int(request.POST.get('warehouse_product_id'))
    count = float(request.POST.get('count'))
    price = int(request.POST.get('price'))
    total = count * price
    obj, created = OrderItem.objects.get_or_create(
        order_id=order_id,
        warehouse_product_id=warehouse_product_id,
        defaults={
            'count': count,
            'price': price,
            'total': total,
        }
    )
    if not created:
        obj.count += count
        obj.price = price
        obj.total = obj.count * obj.price
        obj.save()
    order_item_total = OrderItem.objects.filter(order_id=order_id).aggregate(Sum('total'))
    Order.objects.filter(id=order_id).update(total=order_item_total.get('total__sum', 0))
    return 1


def change_order_status(request, kwargs):
    order_id = kwargs['pk']
    order_status = request.POST.get('status')
    order_instance = Order.objects.get(id=order_id)
    invoice_provider_instance = InvoiceProvider.objects.all()[:1].get()

    if order_status == 'accepted':
        order_items = OrderItem.objects.filter(order=order_instance) \
            .values_list('id', 'count',
                         'price',
                         'warehouse_product__product__price',
                         'warehouse_product_id',
                         'warehouse_product__product__category__bonus')
        order_bonus_item_ids = OrderItem.objects.filter(order=order_instance, item_type=True) \
            .values_list('warehouse_product_id', flat=True)
        invoice = Invoice(order=order_instance, user=request.user, invoice_provider=invoice_provider_instance)
        invoice.save()
        invoice_items = []
        total = 0
        total_without_discount = 0
        for order_item_id, count, show_price, price, wp_id, is_bonus in order_items:
            total += int(count * show_price)
            total_without_discount += int(count * price)
            if is_bonus and wp_id not in order_bonus_item_ids:
                invoice_items.append(InvoiceItem(
                    invoice=invoice,
                    order_item_id=order_item_id,
                    count=count,
                    total=int(count * show_price),
                    item_type='discount',
                ))
            elif wp_id in order_bonus_item_ids:
                invoice_items.append(InvoiceItem(
                    invoice=invoice,
                    order_item_id=order_item_id,
                    count=count,
                    total=int(count * show_price),
                    item_type='bonus',
                ))
            else:
                invoice_items.append(InvoiceItem(
                    invoice=invoice,
                    order_item_id=order_item_id,
                    count=count,
                    total=int(count * show_price),
                    item_type='simple',
                ))
        InvoiceItem.objects.bulk_create(invoice_items)
        invoice.total = total
        invoice.total_without_discount = total_without_discount
        invoice.save()
    if order_status == 'delivered':
        order_items = OrderItem.objects.filter(order=order_instance) \
            .values_list('count', 'warehouse_product_id', )
        invoice_instance = Invoice.objects.get(order_id=order_instance)
        warehouse_products_ids = (elem[1] for elem in order_items)
        warehouse_products = WarehouseProduct.objects.filter(pk__in=warehouse_products_ids).in_bulk()
        for count, warehouse_product_id in order_items:
            warehouse_products[warehouse_product_id].count -= float(count)
        WarehouseProduct.objects.bulk_update(warehouse_products.values(), ['count'])
        counter_party_balance_outcome(order_instance.counter_party_id, invoice_instance.total)
    order_instance.status = order_status
    order_instance.save()
    return 1


def delete_order_item(request, kwargs):
    order_id = kwargs['pk']
    warehouse_product_id = int(request.POST.get('warehouse_product_id'))
    OrderItem.objects.filter(id=warehouse_product_id).delete()
    order_item_total = OrderItem.objects.filter(order_id=order_id).aggregate(Sum('total'))
    if order_item_total.get('total__sum'):
        Order.objects.filter(id=order_id).update(total=order_item_total.get('total__sum', 0))
    else:
        Order.objects.filter(id=order_id).update(total=0)
    return 1


def add_retail_order_item(request, kwargs):
    order_id = kwargs['pk']
    warehouse_product_id = int(request.POST.get('warehouse_product_id'))
    count = float(request.POST.get('count'))
    price = int(request.POST.get('price'))
    total = count * price
    obj, created = RetailOrderItem.objects.get_or_create(
        order_id=order_id,
        warehouse_product_id=warehouse_product_id,
        defaults={
            'count': count,
            'price': price,
            'total': total,
        }
    )
    if not created:
        obj.count += count
        obj.price = price
        obj.total = obj.count * obj.price
        obj.save()
    order_item_total = RetailOrderItem.objects.filter(order_id=order_id).aggregate(Sum('total'))
    RetailOrder.objects.filter(id=order_id).update(total=order_item_total.get('total__sum', 0))
    return 1


def retail_order_payment(request, kwargs):
    retail_order_id = int(request.POST.get('retail_order_id', kwargs.get('pk')))
    retail_order_instance = RetailOrder.objects.get(pk=retail_order_id)
    payment_type = request.POST.get('payment_type')
    payment_method = request.POST.get('payment_method')
    amount = int(request.POST.get('amount'))
    outlay, created = Outlay.objects.get_or_create(outcat='retail_order',
                                                   title='Выплата для погашения долга')
    if payment_type == 'income':
        payment_income_action('retail_order', retail_order_id, outlay, payment_method, amount, request.user)
        client = retail_order_instance.client
        client.balance += amount
        client.save()
    elif payment_type == 'outcome':
        payment_outcome_action('retail_order', retail_order_id, outlay, payment_method, amount, request.user)
        client = retail_order_instance.client
        client.balance -= amount
        client.save()
    return 1


def change_retail_order_status(request, kwargs):
    order_id = kwargs['pk']
    order_status = request.POST.get('status')
    order_instance = RetailOrder.objects.get(id=order_id)

    if order_status == 'completed':
        order_items = RetailOrderItem.objects.filter(order=order_instance).in_bulk()
        for order_item_id, order_item_model in order_items.items():
            wp = order_item_model.warehouse_product
            wp.count -= order_item_model.count
            wp.save()
        client = order_instance.client
        client.balance -= order_instance.total
        client.save()
    elif order_status == 'rejected':
        order_items = RetailOrderItem.objects.filter(order=order_instance).in_bulk()
        for order_item_id, order_item_model in order_items.items():
            wp = order_item_model.warehouse_product
            wp.count += order_item_model.count
            wp.save()
        client = order_instance.client
        client.balance += order_instance.total
        client.save()
    order_instance.status = order_status
    order_instance.save()
    return 1


def delete_retail_order_item(request, kwargs):
    order_id = kwargs['pk']
    warehouse_product_id = int(request.POST.get('warehouse_product_id'))
    RetailOrderItem.objects.filter(id=warehouse_product_id).delete()
    order_item_total = OrderItem.objects.filter(order_id=order_id).aggregate(Sum('total'))
    if order_item_total.get('total__sum'):
        Order.objects.filter(id=order_id).update(total=order_item_total.get('total__sum', 0))
    else:
        Order.objects.filter(id=order_id).update(total=0)
    return 1


def change_invoice_item_counts(request, kwargs):
    order_id = kwargs['pk']
    invoice_items = InvoiceItem.objects \
        .select_related('order_item',
                        'order_item__warehouse_product',
                        'order_item__warehouse_product__product',
                        'order_item__warehouse_product__product__category'
                        ).filter(invoice__order_id=order_id).in_bulk()
    invoice_total = 0
    invoice_total_without_discount = 0
    invoice_provider = request.POST.get('invoice_provider')
    for invoice_item_id, invoice_item_model in invoice_items.items():
        price = invoice_item_model.total_without_discount / invoice_item_model.count
        show_price = invoice_item_model.order_item.price
        invoice_item_model.count = float(request.POST.get(f'invoice_item_{invoice_item_id}_count').replace(',', '.'))
        invoice_item_model.total_without_discount = invoice_item_model.count * show_price
        invoice_item_model.total = invoice_item_model.count * show_price
        if invoice_item_model.item_type == 'discount':
            invoice_item_model.total -= invoice_item_model.total * (float(request.POST.get('invoice_discount')) / 100)
        invoice_total += invoice_item_model.total
        invoice_total_without_discount += invoice_item_model.total_without_discount
    InvoiceItem.objects.bulk_update(invoice_items.values(), fields=('count', 'total', 'total_without_discount'))
    invoice = Invoice.objects.get(order_id=order_id)
    invoice_total -= invoice_total * (float(request.POST.get('invoice_discount')) / 100)
    invoice.total = invoice_total
    if request.POST.get('invoice_deliver'):
        invoice.deliver_id = request.POST.get('invoice_deliver')
        invoice.status = 'during_delivery'
    else:
        invoice.status = 'created'
    invoice.discount = request.POST.get('invoice_discount')
    invoice.invoice_provider_id = invoice_provider
    invoice.save()
    return 1


def write_off_return_items(request, kwargs):
    return_item_id = request.POST.get('return_item_id')
    return_item_write_off = float(request.POST.get('write_off'))
    return_item_obj = ReturnItem.objects.get(pk=return_item_id)
    return_item_obj.status = 'write-off'
    return_item_obj.write_off_count = return_item_write_off
    return_item_obj.returned_count = return_item_obj.count - return_item_write_off
    wp = return_item_obj.warehouse_product
    wp.count += return_item_obj.count - return_item_write_off
    return_item_obj.save()
    wp.save()
    return 1


def reject_return_items(request, kwargs):
    return_item_id = request.POST.get('return_item_id')
    return_item_obj = ReturnItem.objects.get(pk=return_item_id)
    if return_item_obj.status == 'write-off':
        return_item_obj.status = 'rejected'
        wp = return_item_obj.warehouse_product
        wp.count -= return_item_obj.returned_count
    elif return_item_obj.status == 'rejected':
        return_item_obj.status = 'acceptance'
    else:
        return_item_obj.status = 'rejected'
    return_item_obj.save()
    return 1


def return_items_acceptance(request, kwargs):
    order_id = kwargs['pk']
    order_obj = Order.objects.get(pk=order_id)
    return_items = ReturnItem.objects.filter(counter_party_id=order_obj.counter_party_id, status='created').in_bulk()
    for return_item_id, return_item_model in return_items.items():
        return_item_model.count = float(request.POST.get(f'return_item_{return_item_id}_count').replace(',', '.'))
        return_item_model.price = int(request.POST.get(f'return_item_{return_item_id}_price'))
        return_item_model.total = int(request.POST.get(f'return_item_{return_item_id}_total'))
        if request.POST.get(f'return_item_{return_item_id}') == 'on':
            return_item_model.status = 'acceptance'
            cp = return_item_model.counter_party
            cp.balance += return_item_model.total
            cp.save()
    ReturnItem.objects.bulk_update(return_items.values(), fields=('count', 'price', 'total', 'status'))
    return 1


def order_invoice_excel_download(id_):
    invoice = Invoice.objects.filter(id=int(id_)).select_related('order__counter_party', 'invoice_provider').first()
    invoice_items = InvoiceItem.objects.filter(invoice_id=int(id_)).select_related(
        'order_item__warehouse_product__product') \
        .values('count', 'total', 'total_without_discount', 'order_item__warehouse_product__product__title',
                'order_item__warehouse_product__product__unit_type',
                'order_item__warehouse_product__product__price', 'order_item__warehouse_product__product__show_price')

    book = openpyxl.load_workbook('temp.xlsx')
    booksheet = book.worksheets[0]

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    bold_font = Font(bold=True)
    cell_color = PatternFill(fill_type='solid', start_color="b3b6ba", end_color="b3b6ba")

    booksheet['d4'] = f"Товарная накладная №: {invoice.id} Дата: {invoice.created.date()}"
    booksheet['d4'].font = bold_font

    booksheet['b7'] = f'Поставщик: {invoice.invoice_provider.name}'
    booksheet['b8'] = f"Адрес: {invoice.invoice_provider.address}"
    booksheet['b9'] = f"Телефон: {invoice.invoice_provider.phone_number}"

    bold_font = Font(bold=True, size=9)

    booksheet['i7'] = f"Получатель: {invoice.order.counter_party.full_name}"
    booksheet['i8'] = f"Адрес: {invoice.order.counter_party.address}"
    booksheet['i9'] = f"Телефон: {invoice.order.counter_party.phone_number}"
    booksheet['i11'] = f"Долг: {locale.format('%d', invoice.order.counter_party.balance, grouping=True)}"
    booksheet['i11'].border = border
    booksheet['i11'].fill = cell_color
    booksheet['i11'].font = bold_font

    img_load = os.path.join(settings.BASE_DIR + '/media/' + str(invoice.invoice_provider.logo))
    img = openpyxl.drawing.image.Image(img_load)
    img.height = 50
    img.width = 90
    img.anchor = "c10"
    booksheet.add_image(img)

    booksheet['b13'] = f"№"
    booksheet['b13'].font = bold_font
    booksheet['b13'].border = border
    booksheet['c13'] = f"Наименования товаров"
    booksheet['c13'].font = bold_font
    booksheet['c13'].border = border
    booksheet['d13'] = f"Кол-во"
    booksheet['d13'].font = bold_font
    booksheet['d13'].border = border
    booksheet['e13'] = f"Ед. изм."
    booksheet['e13'].font = bold_font
    booksheet['e13'].border = border
    booksheet['f13'] = f"Цена"
    booksheet['f13'].font = bold_font
    booksheet['f13'].border = border
    booksheet['g13'] = f"Цена со скидкой"
    booksheet['g13'].font = bold_font
    booksheet['g13'].border = border
    booksheet['h13'] = f"Сумма"
    booksheet['h13'].font = bold_font
    booksheet['h13'].border = border
    booksheet['i13'] = f"Сумма со скидкой"
    booksheet['i13'].font = bold_font
    booksheet['i13'].border = border

    counter = 0
    count_total_kg = 0
    count_total_p = 0
    price_total = 0
    price_with_discount_total = 0

    for item in invoice_items:
        price_total += item['order_item__warehouse_product__product__price'] * item['count']
        price_with_discount_total += item['order_item__warehouse_product__product__show_price'] * item['count']

        if item['order_item__warehouse_product__product__unit_type'] == 'piece':
            count_total_p += item['count']
            item['order_item__warehouse_product__product__unit_type'] = 'Шт'
        else:
            count_total_kg += item['count']
            item['order_item__warehouse_product__product__unit_type'] = 'Кг'
        counter += 1
        booksheet[f'B{13 + counter}'] = counter
        booksheet[f'B{13 + counter}'].border = border
        booksheet[f'C{13 + counter}'] = item['order_item__warehouse_product__product__title']
        booksheet[f'C{13 + counter}'].border = border
        booksheet[f'D{13 + counter}'] = item['count']
        booksheet[f'D{13 + counter}'].border = border
        booksheet[f'E{13 + counter}'] = item['order_item__warehouse_product__product__unit_type']
        booksheet[f'E{13 + counter}'].border = border
        booksheet[f'F{13 + counter}'] = locale.format('%d', int(
            item['order_item__warehouse_product__product__price']
        ), grouping=True)
        booksheet[f'F{13 + counter}'].border = border
        booksheet[f'G{13 + counter}'] = locale.format('%d', int(
            item['order_item__warehouse_product__product__show_price']
        ), grouping=True)
        booksheet[f'G{13 + counter}'].border = border
        booksheet[f'H{13 + counter}'] = locale.format('%d', int(item['total_without_discount']), grouping=True)
        booksheet[f'H{13 + counter}'].border = border
        booksheet[f'I{13 + counter}'] = locale.format('%d', int(item['total']), grouping=True)
        booksheet[f'I{13 + counter}'].border = border

    booksheet[f"B{14 + counter}"].border = border
    booksheet[f"B{14 + counter}"].fill = cell_color
    booksheet[f"E{14 + counter}"].border = border
    booksheet[f"E{14 + counter}"].fill = cell_color
    booksheet[f"F{14 + counter}"].border = border
    booksheet[f"F{14 + counter}"].fill = cell_color
    booksheet[f"G{14 + counter}"].border = border
    booksheet[f"G{14 + counter}"].fill = cell_color
    booksheet[f"C{14 + counter}"] = "Итого:"
    booksheet[f"C{14 + counter}"].fill = cell_color
    booksheet[f"C{14 + counter}"].border = border
    booksheet[f"C{14 + counter}"].font = bold_font
    booksheet[f"D{14 + counter}"] = f"{count_total_kg} кг/ {count_total_p} шт"
    booksheet[f"D{14 + counter}"].border = border
    booksheet[f"D{14 + counter}"].fill = cell_color
    booksheet[f"D{14 + counter}"].font = bold_font
    booksheet[f"H{14 + counter}"] = f"{locale.format('%d', int(price_total), grouping=True)}"
    booksheet[f"H{14 + counter}"].border = border
    booksheet[f"H{14 + counter}"].fill = cell_color
    booksheet[f"H{14 + counter}"].font = bold_font
    booksheet[f"I{14 + counter}"] = f"{locale.format('%d', int(price_with_discount_total), grouping=True)}"
    booksheet[f"I{14 + counter}"].border = border
    booksheet[f"I{14 + counter}"].fill = cell_color
    booksheet[f"I{14 + counter}"].font = bold_font

    booksheet[f"B{16 + counter}"] = "Долг__________________________"
    booksheet[f"B{16 + counter}"].font = bold_font
    booksheet[f"G{16 + counter}"] = "Расчеть___________________________"
    booksheet[f"G{16 + counter}"].font = bold_font

    booksheet[f"B{18 + counter}"].border = border
    booksheet[f"C{18 + counter}"].border = border
    booksheet[f"D{18 + counter}"].border = border
    booksheet[f"E{18 + counter}"].border = border
    booksheet[f"F{18 + counter}"].border = border
    booksheet[f"G{18 + counter}"].border = border
    booksheet[f"H{18 + counter}"].border = border
    booksheet[f"I{18 + counter}"].border = border

    response = HttpResponse(content=save_virtual_workbook(book),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.id}.xlsx"'

    return response

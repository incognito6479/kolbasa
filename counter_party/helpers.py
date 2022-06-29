from counter_party.models import CounterParty
from django.http import HttpResponse
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.writer.excel import save_virtual_workbook
import openpyxl, os
from payment.models import PaymentLog
from counter_party.models import RetailClient
from order.models import RetailOrder, ReturnItem, Invoice
from datetime import datetime
from payment.models import Outlay


def counter_party_balance_income(counter_party_id, amount):
    counter_party = CounterParty.objects.get(pk=counter_party_id)
    counter_party.balance += amount
    counter_party.save()


def counter_party_balance_outcome(counter_party_id, amount):
    counter_party = CounterParty.objects.get(pk=counter_party_id)
    counter_party.balance -= amount
    counter_party.save()


def retail_client_report_excel_download(pk, request):
    retail_client = RetailClient.objects.get(pk=pk)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    payments = PaymentLog.objects.filter(outcat='retail_order',
                                         status='accepted')

    range_text = ''

    if start_date and end_date:
        start_date += ' 00:00:00'
        end_date += ' 23:59:59'
        range_text = f" с {start_date} до {end_date}"
        retail_orders = RetailOrder.objects.filter(client=retail_client,
                                                   status='completed',
                                                   created__range=[
                                                       start_date, end_date
                                                   ])
        payments = payments.filter(
            model_id__in=retail_orders.values_list('id', flat=True)
        )
    else:
        today_start = datetime.today().date().strftime('%Y-%m-%d') + ' 00:00:00'
        today_end = datetime.today().date().strftime('%Y-%m-%d') + ' 23:59:59'
        range_text = ' Сегодня'
        retail_orders = RetailOrder.objects.filter(client=retail_client,
                                                   status='completed',
                                                   created__range=[
                                                       today_start, today_end
                                                   ])
        payments = payments.filter(
            model_id__in=retail_orders.values_list('id', flat=True)
        )
    retail_orders_dicts = retail_orders.values('created', 'id', 'total', 'user__username')
    retail_orders_dicts = ((retail_order['created'], {'title': 'Заказ №' + str(retail_order['id'])
                                                          + ' ' + retail_order['user__username'],
                                                 'created': retail_order['created'],
                                                 'id': retail_order['id'],
                                                 'type': 'retail_order',
                                                 'type_display': 'Расход',
                                                 'amount': retail_order['total']
                                                 }) for retail_order in retail_orders_dicts)
    payments_dicts = payments.values('created', 'id', 'amount', 'user__fullname')
    payments_dicts = ((
        payments_model['created'], {'title': 'Оплата №' + str(payments_model['id'])
                                             + ' ' + payments_model[
                                                 'user__fullname'],
                                    'created': payments_model['created'],
                                    'id': payments_model['id'],
                                    'type': 'payment',
                                    'type_display': 'Приход',
                                    'amount': payments_model['amount']}
    ) for payments_model in payments_dicts)
    report_dict = list(retail_orders_dicts) + list(payments_dicts)
    report_dict.sort()
    report_dict = {created: model for created, model in report_dict}
    # print(report_dict.values())

    book = openpyxl.load_workbook('client_report.xlsx')
    booksheet = book.worksheets[0]

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    bold_font = Font(bold=True)
    cell_color = PatternFill(fill_type='solid', start_color="b3b6ba", end_color="b3b6ba")

    booksheet['c4'] = f"Акт сверка ({retail_client.full_name} | {retail_client.phone_number} | {retail_client.balance})"
    booksheet['c4'].value += range_text

    booksheet['b6'] = "Адрес клиента"
    booksheet['b6'].font = bold_font
    booksheet['b7'] = "Комментарии о клиенте"
    booksheet['b7'].font = bold_font
    booksheet['b8'] = "Баланс клиента"
    booksheet['b8'].font = bold_font

    booksheet['d6'] = f"{retail_client.address}"
    booksheet['d7'] = f"{retail_client.content}"
    booksheet['d8'] = f"{retail_client.balance} сум на {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    
    bold_font = Font(bold=True, size=9)

    booksheet['a10'] = "№"
    booksheet['a10'].font = bold_font
    booksheet['a10'].border = border
    booksheet['b10'] = "ДАТА"
    booksheet['b10'].font = bold_font
    booksheet['b10'].border = border
    booksheet['c10'] = "ПРИЧИНА"
    booksheet['c10'].font = bold_font
    booksheet['c10'].border = border
    booksheet['d10'] = "ТИП"
    booksheet['d10'].font = bold_font
    booksheet['d10'].border = border
    booksheet['e10'] = "СУММА"
    booksheet['e10'].font = bold_font
    booksheet['e10'].border = border
    
    counter = 0
    for report in report_dict.values():
        counter += 1
        booksheet[f'A{10 + counter}'] = counter
        booksheet[f'A{10 + counter}'].border = border
        booksheet[f'B{10 + counter}'] = report['created'].strftime('%d.%m.%Y %H:%M')
        booksheet[f'B{10 + counter}'].border = border
        booksheet[f'C{10 + counter}'] = report['title']
        booksheet[f'C{10 + counter}'].border = border
        booksheet[f'D{10 + counter}'] = report['type_display']
        booksheet[f'D{10 + counter}'].border = border
        booksheet[f'E{10 + counter}'] = report['amount']
        booksheet[f'E{10 + counter}'].border = border
    if not report_dict.values():
        booksheet[f'A11'].border = border
        booksheet[f'B11'].border = border
        booksheet[f'C11'] = 'Пусто' 
        booksheet[f'C11'].border = border
        booksheet[f'D11'].border = border
        booksheet[f'E11'].border = border

    response = HttpResponse(content=save_virtual_workbook(book),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="client_report_{retail_client.id}.xlsx"'

    return response


def counter_party_report_excel_download(pk, request):
    counter_party = CounterParty.objects.get(pk=pk)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    outlay, created = Outlay.objects.get_or_create(outcat='counter_party',
                                                   title='Выплата для погашения долга')
    payments = PaymentLog.objects.filter(outcat='counter_party',
                                         outlay=outlay,
                                         status='accepted',
                                         model_id=counter_party.id)

    range_text = ''

    if start_date and end_date:
        start_date += ' 00:00:00'
        end_date += ' 23:59:59'
        range_text = f" с {start_date} до {end_date}"
        invoices = Invoice.objects.filter(order__counter_party=counter_party,
                                          status='delivered',
                                          created__range=[
                                              start_date, end_date
                                          ])

        return_items = ReturnItem.objects.filter(counter_party=counter_party,
                                                 status__in=['acceptance', 'write-off'],
                                                 created__range=[
                                                     start_date, end_date
                                                 ])
        payments = payments.filter(
            created__range=[
                start_date, end_date
            ])
    else:
        today_start = datetime.today().date().strftime('%Y-%m-%d') + ' 00:00:00'
        today_end = datetime.today().date().strftime('%Y-%m-%d') + ' 23:59:59'
        range_text = ' Сегодня'
        invoices = Invoice.objects.filter(order__counter_party=counter_party,
                                          status='delivered',
                                          created__range=[
                                              today_start, today_end
                                          ])
        return_items = ReturnItem.objects.filter(counter_party=counter_party,
                                                 status__in=['acceptance', 'write-off'],
                                                 created__range=[
                                                     today_start, today_end
                                                 ])
        payments = payments.filter(
            created__range=[
                today_start, today_end
            ])
    invoices_dicts = invoices.values('created', 'order_id', 'id', 'total', 'deliver__user__fullname')
    invoices_dicts = ((invoice_model['created'], {'title': 'Заказ №' + str(invoice_model['id'])
                                                           + ' ' + invoice_model['deliver__user__fullname'],
                                                  'created': invoice_model['created'],
                                                  'id': invoice_model['order_id'],
                                                  'type': 'invoice',
                                                  'type_display': 'Расход',
                                                  'amount': invoice_model['total']
                                                  }) for invoice_model in invoices_dicts)
    return_items_dicts = return_items.values('created', 'id', 'total', 'deliver__user__fullname')
    return_items_dicts = ((
        return_item_model['created'], {'title': 'Возврат №' + str(return_item_model['id'])
                                                + ' ' + return_item_model[
                                                    'deliver__user__fullname'],
                                       'created': return_item_model['created'],
                                       'id': return_item_model['id'],
                                       'type': 'return_item',
                                       'type_display': 'Приход',
                                       'amount': return_item_model['total']}
    ) for return_item_model in return_items_dicts)
    payments_dicts = payments.values('created', 'id', 'amount', 'user__fullname')
    payments_dicts = ((
        payments_model['created'], {'title': 'Оплата №' + str(payments_model['id'])
                                             + ' ' + payments_model[
                                                 'user__fullname'],
                                    'created': payments_model['created'],
                                    'id': payments_model['id'],
                                    'type': 'payment',
                                    'type_display': 'Приход',
                                    'amount': payments_model['amount']}
    ) for payments_model in payments_dicts)
    report_dict = list(invoices_dicts) + list(return_items_dicts) + list(payments_dicts)
    report_dict.sort()
    report_dict = {created: model for created, model in report_dict}

    book = openpyxl.load_workbook('client_report.xlsx')
    booksheet = book.worksheets[0]

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    bold_font = Font(bold=True)
    cell_color = PatternFill(fill_type='solid', start_color="b3b6ba", end_color="b3b6ba")

    booksheet['c4'] = f"Акт сверка ({counter_party.full_name} | {counter_party.phone_number})"
    booksheet['c4'].value += range_text
    booksheet['c4'].font = bold_font

    booksheet['b6'] = "Адрес контрагента"
    booksheet['b6'].font = bold_font
    booksheet['b7'] = "Регион"
    booksheet['b7'].font = bold_font
    booksheet['b8'] = "Комментарии о контрагента"
    booksheet['b8'].font = bold_font
    booksheet['b9'] = "Баланс контрагента"
    booksheet['b9'].font = bold_font

    booksheet['d6'] = f"{counter_party.address}"
    booksheet['d7'] = f"{counter_party.get_region_display()}"
    booksheet['d8'] = f"{counter_party.content}"
    booksheet['d9'] = f"{counter_party.balance} на {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    
    bold_font = Font(bold=True, size=9)

    booksheet['a11'] = "№"
    booksheet['a11'].font = bold_font
    booksheet['a11'].border = border
    booksheet['b11'] = "ДАТА"
    booksheet['b11'].font = bold_font
    booksheet['b11'].border = border
    booksheet['c11'] = "ПРИЧИНА"
    booksheet['c11'].font = bold_font
    booksheet['c11'].border = border
    booksheet['d11'] = "ТИП"
    booksheet['d11'].font = bold_font
    booksheet['d11'].border = border
    booksheet['e11'] = "СУММА"
    booksheet['e11'].font = bold_font
    booksheet['e11'].border = border
    
    counter = 0
    for report in report_dict.values():
        counter += 1
        booksheet[f'A{11 + counter}'] = counter
        booksheet[f'A{11 + counter}'].border = border
        booksheet[f'B{11 + counter}'] = report['created'].strftime('%d.%m.%Y %H:%M')
        booksheet[f'B{11 + counter}'].border = border
        booksheet[f'C{11 + counter}'] = report['title']
        booksheet[f'C{11 + counter}'].border = border
        booksheet[f'D{11 + counter}'] = report['type_display']
        booksheet[f'D{11 + counter}'].border = border
        booksheet[f'E{11 + counter}'] = report['amount']
        booksheet[f'E{11 + counter}'].border = border
    if not report_dict.values():
        booksheet[f'A12'].border = border
        booksheet[f'B12'].border = border
        booksheet[f'C12'] = 'Пусто' 
        booksheet[f'C12'].border = border
        booksheet[f'D12'].border = border
        booksheet[f'E12'].border = border
        

    response = HttpResponse(content=save_virtual_workbook(book),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="counter_party_{counter_party.id}.xlsx"'

    return response